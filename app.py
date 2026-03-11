from __future__ import annotations

import csv
import io
import os
import sqlite3
from contextlib import closing
from datetime import datetime
from pathlib import Path
from typing import Iterable

from flask import Flask, flash, g, redirect, render_template, request, url_for

BASE_DIR = Path(__file__).resolve().parent
DATABASE = BASE_DIR / "assets.db"

app = Flask(__name__)
app.config["SECRET_KEY"] = os.getenv("ASSET_APP_SECRET", "dev-secret-key")


# ---------- DB helpers ----------
def get_db() -> sqlite3.Connection:
    if "db" not in g:
        g.db = sqlite3.connect(DATABASE)
        g.db.row_factory = sqlite3.Row
    return g.db


@app.teardown_appcontext
def close_db(_: object) -> None:
    db = g.pop("db", None)
    if db is not None:
        db.close()


def init_db() -> None:
    schema = """
    CREATE TABLE IF NOT EXISTS employees (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        name TEXT NOT NULL UNIQUE,
        status TEXT NOT NULL DEFAULT 'active',
        created_at TEXT NOT NULL
    );

    CREATE TABLE IF NOT EXISTS assets (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        sn TEXT NOT NULL UNIQUE,
        model TEXT NOT NULL,
        password TEXT,
        status TEXT NOT NULL DEFAULT 'idle',
        employee_id INTEGER,
        created_at TEXT NOT NULL,
        updated_at TEXT NOT NULL,
        FOREIGN KEY(employee_id) REFERENCES employees(id)
    );
    """
    db = sqlite3.connect(DATABASE)
    with closing(db.cursor()) as cursor:
        cursor.executescript(schema)
    db.commit()
    db.close()


# ---------- Core business logic ----------
def now_text() -> str:
    return datetime.now().strftime("%Y-%m-%d %H:%M:%S")


def release_assets_for_departed_employee(employee_id: int) -> None:
    db = get_db()
    db.execute(
        """
        UPDATE assets
        SET employee_id = NULL,
            status = 'idle',
            updated_at = ?
        WHERE employee_id = ?
        """,
        (now_text(), employee_id),
    )
    db.commit()


def upsert_asset(sn: str, model: str, password: str, employee_name: str | None) -> tuple[bool, str]:
    db = get_db()
    sn = sn.strip()
    model = model.strip()
    password = (password or "").strip()
    employee_name = (employee_name or "").strip()

    if not sn or not model:
        return False, "SN 和型号不能为空"

    employee_id = None
    status = "idle"
    if employee_name:
        existing = db.execute("SELECT id FROM employees WHERE name = ?", (employee_name,)).fetchone()
        if existing is None:
            db.execute(
                "INSERT INTO employees(name, status, created_at) VALUES (?, 'active', ?)",
                (employee_name, now_text()),
            )
            employee_id = db.execute("SELECT last_insert_rowid() AS id").fetchone()["id"]
        else:
            employee_id = existing["id"]
        status = "in_use"

    existing_asset = db.execute("SELECT id FROM assets WHERE sn = ?", (sn,)).fetchone()
    if existing_asset:
        db.execute(
            """
            UPDATE assets
            SET model = ?,
                password = ?,
                status = ?,
                employee_id = ?,
                updated_at = ?
            WHERE id = ?
            """,
            (model, password, status, employee_id, now_text(), existing_asset["id"]),
        )
        db.commit()
        return True, f"已更新资产 {sn}"

    db.execute(
        """
        INSERT INTO assets(sn, model, password, status, employee_id, created_at, updated_at)
        VALUES (?, ?, ?, ?, ?, ?, ?)
        """,
        (sn, model, password, status, employee_id, now_text(), now_text()),
    )
    db.commit()
    return True, f"已新增资产 {sn}"


def parse_rows_from_upload(filename: str, file_bytes: bytes) -> Iterable[dict[str, str]]:
    suffix = Path(filename.lower()).suffix

    if suffix == ".csv":
        content = file_bytes.decode("utf-8-sig")
        reader = csv.DictReader(io.StringIO(content))
        for row in reader:
            yield {k.strip(): str(v or "").strip() for k, v in row.items() if k}
        return

    if suffix == ".xlsx":
        try:
            from openpyxl import load_workbook
        except ModuleNotFoundError as exc:
            raise RuntimeError("导入 xlsx 需要先安装 openpyxl：pip install openpyxl") from exc

        wb = load_workbook(io.BytesIO(file_bytes), data_only=True)
        ws = wb.active
        header = [str(c.value).strip() if c.value is not None else "" for c in ws[1]]
        for row in ws.iter_rows(min_row=2, values_only=True):
            item: dict[str, str] = {}
            for idx, val in enumerate(row):
                key = header[idx] if idx < len(header) else f"col_{idx}"
                item[key] = "" if val is None else str(val).strip()
            yield item
        return

    raise RuntimeError("仅支持 .csv 或 .xlsx 文件")


def normalize_import_row(row: dict[str, str]) -> tuple[str, str, str, str]:
    keys = {k.lower(): v for k, v in row.items()}

    def first_of(*candidates: str) -> str:
        for c in candidates:
            if c in row and row[c]:
                return row[c]
            cl = c.lower()
            if cl in keys and keys[cl]:
                return keys[cl]
        return ""

    employee = first_of("所属员工", "员工", "employee", "owner", "使用人")
    sn = first_of("资产SN", "SN", "sn", "资产编号", "序列号")
    model = first_of("资产型号", "型号", "model")
    password = first_of("电脑密码", "password", "密码")
    return employee, sn, model, password


# ---------- Routes ----------
@app.route("/")
def index():
    db = get_db()
    assets = db.execute(
        """
        SELECT a.*, e.name AS employee_name
        FROM assets a
        LEFT JOIN employees e ON a.employee_id = e.id
        ORDER BY a.updated_at DESC
        """
    ).fetchall()
    employees = db.execute("SELECT * FROM employees ORDER BY name").fetchall()
    return render_template("index.html", assets=assets, employees=employees)


@app.post("/employees")
def create_employee():
    name = request.form.get("name", "").strip()
    if not name:
        flash("员工名不能为空", "danger")
        return redirect(url_for("index"))

    db = get_db()
    try:
        db.execute(
            "INSERT INTO employees(name, status, created_at) VALUES (?, 'active', ?)",
            (name, now_text()),
        )
        db.commit()
        flash(f"员工 {name} 已添加", "success")
    except sqlite3.IntegrityError:
        flash(f"员工 {name} 已存在", "warning")
    return redirect(url_for("index"))


@app.post("/employees/<int:employee_id>/depart")
def depart_employee(employee_id: int):
    db = get_db()
    employee = db.execute("SELECT * FROM employees WHERE id = ?", (employee_id,)).fetchone()
    if employee is None:
        flash("未找到员工", "danger")
        return redirect(url_for("index"))

    db.execute("UPDATE employees SET status = 'departed' WHERE id = ?", (employee_id,))
    db.commit()
    release_assets_for_departed_employee(employee_id)
    flash(f"员工 {employee['name']} 已离职，名下资产已转为闲置", "success")
    return redirect(url_for("index"))


@app.post("/assets")
def create_asset():
    employee = request.form.get("employee", "")
    sn = request.form.get("sn", "")
    model = request.form.get("model", "")
    password = request.form.get("password", "")
    ok, msg = upsert_asset(sn=sn, model=model, password=password, employee_name=employee)
    flash(msg, "success" if ok else "danger")
    return redirect(url_for("index"))


@app.post("/assets/<int:asset_id>/assign")
def assign_asset(asset_id: int):
    employee_id = int(request.form.get("employee_id", "0"))
    db = get_db()
    asset = db.execute("SELECT * FROM assets WHERE id = ?", (asset_id,)).fetchone()
    employee = db.execute(
        "SELECT * FROM employees WHERE id = ? AND status = 'active'", (employee_id,)
    ).fetchone()

    if asset is None or employee is None:
        flash("资产或员工不存在", "danger")
        return redirect(url_for("index"))

    db.execute(
        "UPDATE assets SET employee_id = ?, status = 'in_use', updated_at = ? WHERE id = ?",
        (employee_id, now_text(), asset_id),
    )
    db.commit()
    flash(f"资产 {asset['sn']} 已分配给 {employee['name']}", "success")
    return redirect(url_for("index"))


@app.post("/assets/<int:asset_id>/idle")
def mark_idle(asset_id: int):
    db = get_db()
    db.execute(
        "UPDATE assets SET employee_id = NULL, status = 'idle', updated_at = ? WHERE id = ?",
        (now_text(), asset_id),
    )
    db.commit()
    flash("资产已转为闲置", "success")
    return redirect(url_for("index"))


@app.post("/assets/import")
def import_assets():
    upload = request.files.get("file")
    if upload is None or not upload.filename:
        flash("请选择要导入的文件", "danger")
        return redirect(url_for("index"))

    try:
        rows = list(parse_rows_from_upload(upload.filename, upload.read()))
    except RuntimeError as exc:
        flash(str(exc), "danger")
        return redirect(url_for("index"))

    success_count = 0
    for row in rows:
        employee, sn, model, password = normalize_import_row(row)
        ok, _ = upsert_asset(sn=sn, model=model, password=password, employee_name=employee)
        if ok:
            success_count += 1

    flash(f"导入完成：成功处理 {success_count} 条", "success")
    return redirect(url_for("index"))


if __name__ == "__main__":
    init_db()
    app.run(host="0.0.0.0", port=5000, debug=True)
